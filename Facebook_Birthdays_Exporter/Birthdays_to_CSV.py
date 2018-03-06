import csv
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# Get directory of excel file to process and load it
xlDir = input("Enter the path of the Excel File to be processed: ")
wb = Workbook()
wb = load_workbook(xlDir)

# ws = the active worksheet
ws = wb.active

# Cell, counter and other temporary variables for the while loop
i = 0
c = 'A1'
name = 'Rick Astley'
date = '30 February'
cContent = ''

# Create the output file
outFile = open('Birthdays.csv', 'w+')
outFile.close()

# Open the output csv file, examine all cells in column, perform operations, write output to csv
with open('Birthdays.csv', 'w', newline='', encoding='utf-8') as f:
    csvWriter = csv.writer(f, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    while i < ws.max_row :
        i = i+1
        c = 'A' + str(i)

        # Temporary string var to hold cell content
        cContent = ws[c].value

        # Read Date
        if (cContent[0:17] == 'DTSTART;VALUE=DATE:') :
            # Exclude: DTSTART;VALUE=DATE:yyyy
            date = cContent[21:]
            if date[0:2] == '01':
                date = " January " + date[2:]
            elif date[0:2] == '02':
                date = " February " + date[2:]
            elif date[0:2] == '03':
                date = " March " + date[2:]
            elif date[0:2] == '04':
                date = " April " + date[2:]
            elif date[0:2] == '05':
                date = " May " + date[2:]
            elif date[0:2] == '06':
                date = " June " + date[2:]
            elif date[0:2] == '07':
                date = " July " + date[2:]
            elif date[0:2] == '08':
                date = " August " + date[2:]
            elif date[0:2] == '09':
                date = " September " + date[2:]
            elif date[0:2] == '10':
                date = " October " + date[2:]
            elif date[0:2] == '11':
                date = " November " + date[2:]
            elif date[0:2] == '12':
                date = " December " + date[2:]

        # Read name
        if (cContent[0:8] == 'SUMMARY:') :
            # Exclude: SUMMARY:
            name = cContent[8:]
            # Exclude: 's birthday
            name = name[:-11]

        # If you find this line, it's time to write to csv
        if (cContent == "END:VEVENT") :
            csvWriter.writerow([name, date])
